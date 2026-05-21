"""
Read-only content audit for atelier site.

Usage (from project root):
  python scripts/site/audit_content.py
  python scripts/site/audit_content.py --suggest
  python scripts/site/audit_content.py --out reports/content-audit-r2.md
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent
WIKI_DIR = ROOT / "Wiki"
BLOG_DIR = ROOT / "Blog"
XLSX_PATH = ROOT / "zhita_settings.xlsx"
FRAMEWORK_MANIFEST = BLOG_DIR / "framework-guides" / "manifest.json"

IMG_REF_RE = re.compile(r"!\[[^\]]*\]\(([^)]+)\)")
DETAILS_RE = re.compile(r"<details>[\s\S]*?Relevant\s+source\s+files", re.I)


def audit_xlsx(lines: list[str], suggest: bool) -> None:
    lines.append("## zhita_settings.xlsx")
    if not XLSX_PATH.is_file():
        lines.append("- **缺失**: zhita_settings.xlsx 不存在")
        return
    try:
        from openpyxl import load_workbook

        wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            data_rows = [
                r for r in rows[1:]
                if r and any(c is not None and str(c).strip() for c in r)
            ]
            lines.append(f"- 表 `{name}`: {len(data_rows)} 条数据行")
        wb.close()
    except Exception as exc:
        lines.append(f"- **错误**: 无法读取 xlsx — {exc}")
        return

    if FRAMEWORK_MANIFEST.is_file():
        try:
            manifest = json.loads(FRAMEWORK_MANIFEST.read_text(encoding="utf-8"))
            posts = manifest.get("posts", [])
            draft = sum(1 for p in posts if p.get("status") != "published")
            lines.append(f"- Framework manifest: {len(posts)} 篇，连载稿 {draft} 篇")
        except json.JSONDecodeError as exc:
            lines.append(f"- Framework manifest JSON 错误: {exc}")

    if suggest:
        lines.append("\n### 建议")
        lines.append("- 更新 xlsx 后执行 `from site_data import clear_site_data_cache; clear_site_data_cache()` 或重启服务")
        lines.append("- Wiki 源文件含 details 时运行 `python scripts/site/clean_wiki_sources.py --write`")


def audit_wiki(lines: list[str]) -> None:
    lines.append("\n## Wiki")
    if not WIKI_DIR.is_dir():
        lines.append("- Wiki 目录不存在")
        return
    for slug_dir in sorted(WIKI_DIR.iterdir()):
        if not slug_dir.is_dir():
            continue
        slug = slug_dir.name
        issues: list[str] = []
        for md in slug_dir.glob("*.md"):
            if "_source" in md.parts or md.name == "README.md":
                continue
            text = md.read_text(encoding="utf-8", errors="replace")
            if DETAILS_RE.search(text):
                issues.append(f"{md.name}: 含 Relevant source files")
            h1 = len(re.findall(r"^#\s+", text, re.M))
            if h1 >= 2:
                issues.append(f"{md.name}: 多个 H1 ({h1})")
            for m in IMG_REF_RE.finditer(text):
                ref = m.group(1)
                if ref.startswith("http"):
                    continue
                local = slug_dir / ref
                if not local.is_file():
                    issues.append(f"{md.name}: 缺图 {ref}")
        if issues:
            lines.append(f"### {slug}")
            for i in issues[:12]:
                lines.append(f"- {i}")
            if len(issues) > 12:
                lines.append(f"- …共 {len(issues)} 项")


def audit_blog(lines: list[str]) -> None:
    lines.append("\n## Blog")
    standalone = BLOG_DIR / "认识简谱" / "index.md"
    if standalone.is_file():
        raw = standalone.read_text(encoding="utf-8", errors="replace")
        for m in IMG_REF_RE.finditer(raw):
            ref = m.group(1)
            if ref.startswith("images/"):
                src = BLOG_DIR / "认识简谱" / ref
                static = ROOT / "static" / "blog" / "jianpu" / ref.replace("images/", "")
                if not src.is_file() and not static.is_file():
                    lines.append(f"- jianpu: 缺图 {ref}")

    if not FRAMEWORK_MANIFEST.is_file():
        return
    try:
        manifest = json.loads(FRAMEWORK_MANIFEST.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        lines.append("- Framework manifest 无法解析")
        return
    for entry in manifest.get("posts", []):
        slug = entry.get("slug", "")
        folder = entry.get("folder", "")
        path = BLOG_DIR / folder / "index.md"
        if not path.is_file():
            lines.append(f"- **缺失**: {slug} — index.md")
            continue
        raw = path.read_text(encoding="utf-8", errors="replace")
        if entry.get("series") == "framework" and "## " not in raw:
            lines.append(f"- {slug}: Framework 稿无 ## 章节")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--suggest", action="store_true")
    parser.add_argument("--out", type=Path, default=None)
    args = parser.parse_args()

    report: list[str] = ["# 内容审计报告（第二轮）", ""]
    audit_xlsx(report, args.suggest)
    audit_wiki(report)
    audit_blog(report)

    text = "\n".join(report) + "\n"
    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(text, encoding="utf-8")
        print(f"Wrote {args.out}")
    else:
        print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
