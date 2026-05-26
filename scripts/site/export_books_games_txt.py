"""Export / import 书籍、游戏 sheets from zhita_settings.xlsx as UTF-8 TSV txt."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = ROOT / "zhita_settings.xlsx"
EXPORT_DIR = ROOT / "data" / "export"

BOOKS_FILE = EXPORT_DIR / "书籍.txt"
GAMES_FILE = EXPORT_DIR / "游戏.txt"

BOOKS_SHEET = "书籍"
GAMES_SHEET = "游戏"
BOOKS_HEADER = ("书名", "作者", "分类")
GAMES_HEADER = ("名称", "系列", "平台")

BOOKS_HELP = """\
# =============================================================================
# 书籍数据（批量编辑用）
# 来源: zhita_settings.xlsx → 工作表「书籍」
# 写回: python scripts/site/export_books_games_txt.py --import-books
# 说明:
#   - 制表符分隔；以 # 开头的行为注释；空行忽略
#   - 下方第一行有效数据必须是表头：书名、作者、分类（Tab 分隔）
#   - 从第二行起每行一条记录；字段内请勿使用制表符
# =============================================================================
"""

GAMES_HELP = """\
# =============================================================================
# 游戏数据（批量编辑用）
# 来源: zhita_settings.xlsx → 工作表「游戏」
# 写回: python scripts/site/export_books_games_txt.py --import-games
# 说明:
#   - 制表符分隔；以 # 开头的行为注释；空行忽略
#   - 下方第一行有效数据必须是表头：名称、系列、平台（Tab 分隔）
#   - 从第二行起每行一条记录；字段内请勿使用制表符
# =============================================================================
"""


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _load_workbook():
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("需要 openpyxl：pip install openpyxl") from e
    if not XLSX_PATH.is_file():
        raise SystemExit(f"未找到 Excel：{XLSX_PATH}")
    return openpyxl


def _read_sheet_rows(sheet_name: str) -> list[list[str]]:
    openpyxl = _load_workbook()
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    out: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        out.append([_cell_str(c) for c in row])
    wb.close()
    return out


def _parse_data_rows(
    rows: list[list[str]], min_cols: int, required_col0: bool = True
) -> list[tuple[str, ...]]:
    if len(rows) < 2:
        return []
    data: list[tuple[str, ...]] = []
    for row in rows[1:]:
        if not row or (required_col0 and not row[0]):
            continue
        cells = list(row[:min_cols])
        while len(cells) < min_cols:
            cells.append("")
        if required_col0 and not cells[0]:
            continue
        data.append(tuple(cells))
    return data


def _rows_to_tsv_lines(rows: list[tuple[str, ...]]) -> list[str]:
    return ["\t".join(c.replace("\t", " ").replace("\r", " ").replace("\n", " ") for c in row) for row in rows]


def _write_txt(path: Path, help_text: str, header: tuple[str, ...], rows: list[tuple[str, ...]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [help_text.rstrip(), ""]
    lines.append("\t".join(header))
    lines.extend(_rows_to_tsv_lines(rows))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")


def _parse_txt(path: Path, expected_header: tuple[str, ...]) -> list[tuple[str, ...]]:
    if not path.is_file():
        raise SystemExit(f"未找到文件：{path}")
    raw = path.read_text(encoding="utf-8-sig")
    lines = raw.splitlines()
    data_lines: list[str] = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        data_lines.append(s)
    if not data_lines:
        return []
    header_parts = data_lines[0].split("\t")
    if tuple(header_parts) != expected_header:
        raise SystemExit(
            f"{path.name} 表头应为：{' | '.join(expected_header)}，"
            f"实际为：{' | '.join(header_parts)}"
        )
    rows: list[tuple[str, ...]] = []
    n = len(expected_header)
    for line in data_lines[1:]:
        parts = line.split("\t")
        while len(parts) < n:
            parts.append("")
        if not parts[0].strip():
            continue
        rows.append(tuple(parts[:n]))
    return rows


def _write_sheet(sheet_name: str, header: tuple[str, ...], rows: list[tuple[str, ...]]) -> None:
    openpyxl = _load_workbook()
    wb = openpyxl.load_workbook(XLSX_PATH)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        ws = wb.create_sheet(sheet_name)
    ws.append(list(header))
    for row in rows:
        ws.append(list(row))
    wb.save(XLSX_PATH)
    wb.close()


def export_books() -> int:
    rows = _parse_data_rows(_read_sheet_rows(BOOKS_SHEET), 3)
    _write_txt(BOOKS_FILE, BOOKS_HELP, BOOKS_HEADER, rows)
    print(f"已导出 {len(rows)} 条书籍 → {BOOKS_FILE}")
    return len(rows)


def export_games() -> int:
    rows = _parse_data_rows(_read_sheet_rows(GAMES_SHEET), 3)
    _write_txt(GAMES_FILE, GAMES_HELP, GAMES_HEADER, rows)
    print(f"已导出 {len(rows)} 条游戏 → {GAMES_FILE}")
    return len(rows)


def import_books() -> int:
    rows = _parse_txt(BOOKS_FILE, BOOKS_HEADER)
    _write_sheet(BOOKS_SHEET, BOOKS_HEADER, rows)
    print(f"已写回 {len(rows)} 条书籍到 {XLSX_PATH}（工作表「{BOOKS_SHEET}」）")
    print("请重启 python main.py 或调用 site_data.clear_site_data_cache() 使站点刷新。")
    return len(rows)


def import_games() -> int:
    rows = _parse_txt(GAMES_FILE, GAMES_HEADER)
    _write_sheet(GAMES_SHEET, GAMES_HEADER, rows)
    print(f"已写回 {len(rows)} 条游戏到 {XLSX_PATH}（工作表「{GAMES_SHEET}」）")
    print("请重启 python main.py 或调用 site_data.clear_site_data_cache() 使站点刷新。")
    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="导出/导入 书籍、游戏 txt（TSV）")
    parser.add_argument(
        "--export-books",
        action="store_true",
        help="从 Excel 导出书籍到 data/export/书籍.txt",
    )
    parser.add_argument(
        "--export-games",
        action="store_true",
        help="从 Excel 导出游戏到 data/export/游戏.txt",
    )
    parser.add_argument(
        "--import-books",
        action="store_true",
        help="将 data/export/书籍.txt 写回 Excel",
    )
    parser.add_argument(
        "--import-games",
        action="store_true",
        help="将 data/export/游戏.txt 写回 Excel",
    )
    args = parser.parse_args()
    any_flag = (
        args.export_books
        or args.export_games
        or args.import_books
        or args.import_games
    )
    if not any_flag:
        export_books()
        export_games()
        return
    if args.export_books:
        export_books()
    if args.export_games:
        export_games()
    if args.import_books:
        import_books()
    if args.import_games:
        import_games()


if __name__ == "__main__":
    main()
