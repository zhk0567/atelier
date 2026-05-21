"""
Ensure certificate images from zhita_settings.xlsx exist under static/uploads/events/.

Usage (from project root):
  pip install requests openpyxl
  python scripts/site/fetch_cert_images.py

Optional: set CERT_SOURCE_DIR to a folder containing the image files by basename.
"""

from __future__ import annotations

import os
import re
import shutil
import sys
from pathlib import Path
from urllib.parse import urlparse

try:
    import requests
except ImportError:
    requests = None

ROOT = Path(__file__).resolve().parent.parent.parent
XLSX = ROOT / "zhita_settings.xlsx"
OUT_DIR = ROOT / "static" / "uploads" / "events"
SHEET = "活动竞赛证书"
USER_AGENT = "atelier-local/1.0 (personal wiki; cert fetch)"


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_cert_paths() -> list[tuple[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"Sheet not found: {SHEET}", file=sys.stderr)
        return []
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    out: list[tuple[str, str]] = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = _cell_str(row[0])
        image = _cell_str(row[1]) if len(row) > 1 else ""
        if name and image:
            out.append((name, image))
    return out


def normalize_dest(image_ref: str) -> Path | None:
    ref = image_ref.strip().replace("\\", "/")
    if not ref:
        return None
    if ref.startswith("http://") or ref.startswith("https://"):
        name = Path(urlparse(ref).path).name
        return OUT_DIR / name if name else None
    if ref.startswith("/static/uploads/events/"):
        return ROOT / ref.lstrip("/").replace("/", os.sep)
    if "/" not in ref:
        return OUT_DIR / ref
    return OUT_DIR / Path(ref).name


def try_copy_from_source(basename: str, dest: Path, source_dir: Path) -> bool:
    for candidate in (source_dir / basename, source_dir / "events" / basename):
        if candidate.is_file():
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(candidate, dest)
            print(f"copy {basename} <- {candidate}")
            return True
    return False


def download_http(url: str, dest: Path) -> bool:
    if not requests:
        print("requests not installed; cannot download URLs", file=sys.stderr)
        return False
    try:
        r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=30)
        r.raise_for_status()
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(r.content)
        print(f"download {dest.name} <- {url}")
        return True
    except Exception as exc:
        print(f"  fail {url}: {exc}")
        return False


def main() -> int:
    if not XLSX.is_file():
        print(f"Missing {XLSX}", file=sys.stderr)
        return 1

    source_dir = Path(os.environ.get("CERT_SOURCE_DIR", "")).expanduser()
    certs = load_cert_paths()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    ok, skip, missing = 0, 0, 0
    for name, image_ref in certs:
        dest = normalize_dest(image_ref)
        if not dest:
            print(f"skip invalid path: {image_ref!r}")
            missing += 1
            continue
        if dest.is_file() and dest.stat().st_size > 0:
            print(f"skip {dest.name} (exists)")
            skip += 1
            continue

        ref = image_ref.strip()
        if ref.startswith("http://") or ref.startswith("https://"):
            if download_http(ref, dest):
                ok += 1
            else:
                missing += 1
            continue

        basename = dest.name
        if source_dir.is_dir() and try_copy_from_source(basename, dest, source_dir):
            ok += 1
            continue

        print(f"missing {basename} — place file in {OUT_DIR} or set CERT_SOURCE_DIR")
        missing += 1

    print(f"\nDone: {ok} added, {skip} already present, {missing} still missing")
    return 0 if missing == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
