"""Load and merge site content from zhita_settings.xlsx."""

from __future__ import annotations

import json
import os
import re
from functools import lru_cache
from pathlib import Path
from urllib.parse import unquote

BASE_DIR = Path(__file__).resolve().parent
XLSX_PATH = BASE_DIR / "zhita_settings.xlsx"
IDENTITY_PATH = BASE_DIR / "site_identity.json"
BLOG_DIR = BASE_DIR / "Blog"

from app.config import (  # noqa: E402
    blog_algorithm_dir_name,
    blog_framework_dir_name,
    blog_hotspot_dir_name,
    algorithm_manifest_path,
    framework_manifest_path,
    hotspot_manifest_path,
)

FRAMEWORK_MANIFEST_PATH = framework_manifest_path()
ALGORITHM_MANIFEST_PATH = algorithm_manifest_path()
_BLOG_FRAMEWORK_DIR = blog_framework_dir_name()
_BLOG_ALGORITHM_DIR = blog_algorithm_dir_name()
_BLOG_HOTSPOT_DIR = blog_hotspot_dir_name()

# Standalone posts outside Framework manifest
BLOG_STANDALONE: list[dict[str, str]] = [
    {
        "slug": "jianpu",
        "folder": "认识简谱",
        "title": "认识简谱",
        "series": "",
        "category": "音乐",
        "summary": "从零读懂简谱：音高、节奏与休止，看懂数字记谱即可哼唱旋律。",
    },
    {
        "slug": "dataviz-ch09",
        "folder": "数据可视化-第九章",
        "title": "数据可视化技术 · 第九章 pyecharts",
        "series": "",
        "category": "课程笔记",
        "summary": "pyecharts 从入门到组合图、主题与 Web 整合，含教材例程、虎扑实战与课后习题全文。",
    },
]


@lru_cache(maxsize=1)
def load_framework_manifest() -> dict:
    if not FRAMEWORK_MANIFEST_PATH.is_file():
        return {"posts": []}
    return json.loads(FRAMEWORK_MANIFEST_PATH.read_text(encoding="utf-8"))


@lru_cache(maxsize=1)
def load_algorithm_manifest() -> dict:
    if not ALGORITHM_MANIFEST_PATH.is_file():
        return {"posts": []}
    return json.loads(ALGORITHM_MANIFEST_PATH.read_text(encoding="utf-8"))


@lru_cache(maxsize=1)
def load_hotspot_manifest() -> dict:
    path = hotspot_manifest_path()
    if not path.is_file():
        return {"posts": []}
    return json.loads(path.read_text(encoding="utf-8"))


def _manifest_entry_to_post(entry: dict, default_dir: str) -> dict[str, str]:
    folder = entry.get("folder") or f"{default_dir}/{entry['slug']}"
    return {
        "slug": entry["slug"],
        "folder": folder,
        "title": entry.get("title", entry["slug"]),
        "series": entry.get("series", ""),
        "category": entry.get("category", ""),
        "stack": entry.get("stack", entry.get("topic_path", "")),
        "topic_path": entry.get("topic_path", ""),
        "guide_tier": entry.get("guide_tier", ""),
        "summary": entry.get("summary", ""),
    }


@lru_cache(maxsize=1)
def load_all_blog_posts() -> list[dict[str, str]]:
    posts: list[dict[str, str]] = [dict(p) for p in BLOG_STANDALONE]
    for entry in load_framework_manifest().get("posts", []):
        if entry.get("status") != "published":
            continue
        posts.append(_manifest_entry_to_post(entry, _BLOG_FRAMEWORK_DIR))
    for entry in load_algorithm_manifest().get("posts", []):
        if entry.get("status") != "published":
            continue
        posts.append(_manifest_entry_to_post(entry, _BLOG_ALGORITHM_DIR))
    for entry in load_hotspot_manifest().get("posts", []):
        if entry.get("status") != "published":
            continue
        posts.append(_manifest_entry_to_post(entry, _BLOG_HOTSPOT_DIR))
    return posts


def blog_url_for_title(title: str) -> str:
    for post in load_all_blog_posts():
        if post.get("title") == title:
            return f"/blog/{post['slug']}"
    return ""


def get_blog_post(slug: str) -> dict[str, str] | None:
    for post in load_all_blog_posts():
        if post.get("slug") == slug:
            return dict(post)
    for entry in load_framework_manifest().get("posts", []):
        if entry.get("slug") == slug:
            return _manifest_entry_to_post(entry, _BLOG_FRAMEWORK_DIR)
    for entry in load_algorithm_manifest().get("posts", []):
        if entry.get("slug") == slug:
            return _manifest_entry_to_post(entry, _BLOG_ALGORITHM_DIR)
    for entry in load_hotspot_manifest().get("posts", []):
        if entry.get("slug") == slug:
            return _manifest_entry_to_post(entry, _BLOG_HOTSPOT_DIR)
    return None


def list_framework_posts(*, include_draft: bool = False) -> list[dict]:
    manifest = load_framework_manifest()
    out: list[dict] = []
    for entry in manifest.get("posts", []):
        if not include_draft and entry.get("status") != "published":
            continue
        out.append(dict(entry))
    return out


def list_algorithm_posts(*, include_draft: bool = False) -> list[dict]:
    manifest = load_algorithm_manifest()
    out: list[dict] = []
    for entry in manifest.get("posts", []):
        if not include_draft and entry.get("status") != "published":
            continue
        out.append(dict(entry))
    return out


def list_hotspot_posts(*, include_draft: bool = False) -> list[dict]:
    manifest = load_hotspot_manifest()
    out: list[dict] = []
    for entry in manifest.get("posts", []):
        if not include_draft and entry.get("status") != "published":
            continue
        out.append(_manifest_entry_to_post(entry, _BLOG_HOTSPOT_DIR))
    return out

_GITHUB_REPO_RE = re.compile(r"github\.com/[^/]+/([^/?#]+)", re.I)

DEFAULT_SITE = {
    "site_name": "zhk",
    "site_title": "zhk · 创作者与工程师",
    "tagline": "用数据与工程，把算法做成能上线的产品",
    "hero_subtitle": "数据科学 · Android / Kotlin · 计算机视觉",
    "bio": (
        "数据科学与大数据技术在读，主攻 Python / Kotlin 与 Android，做 CV、语音与 OCR 从训练到端侧落地。"
        "项目以移动应用与视觉为主；数模省一、美赛 H 奖、华数杯 O 奖，英语六级。"
    ),
    "focus": "Python · CV / OCR · 移动与多模态",
    "availability": "在读学生 · 项目与竞赛导向",
    "ui": {
        "label_about": "关于我",
        "label_skills": "数据入口",
        "label_interests": "爱好",
        "label_tech_tags": "技术栈",
        "browse_lead": "",
        "stat_projects": "应用项目",
        "stat_projects_unit": "个",
        "stat_skills": "数据分类",
        "stat_skills_unit": "类",
        "stat_media": "阅读与作品",
        "stat_media_unit": "部",
        "label_projects": "精选项目",
        "link_all_projects": "项目",
        "label_all_projects_page": "项目",
        "projects_page_lead": "表格「项目应用」与 Wiki 仓库同步；首页为置顶展示。",
        "badge_pinned": "置顶",
        "fact_education": "教育",
        "fact_reading": "阅读",
    },
}


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _github_repo_key(url: str) -> str:
    if not url:
        return ""
    m = _GITHUB_REPO_RE.search(url)
    if not m:
        return ""
    return unquote(m.group(1)).lower()


def _load_identity_json() -> dict:
    if not IDENTITY_PATH.is_file():
        return {"site_name": DEFAULT_SITE["site_name"], "site_title": DEFAULT_SITE["site_title"]}
    data = json.loads(IDENTITY_PATH.read_text(encoding="utf-8"))
    return {
        "site_name": _cell_str(data.get("site_name", DEFAULT_SITE["site_name"])),
        "site_title": _cell_str(data.get("site_title", DEFAULT_SITE["site_title"])),
    }


def _load_sheet_rows(sheet_name: str) -> list[list[object | None]]:
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("需要 openpyxl：pip install openpyxl") from e

    if not XLSX_PATH.is_file():
        return []

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _resolve_project_image_url(url: str) -> str:
    if not url:
        return ""
    ref = url.strip().replace("\\", "/")
    if ref.startswith("http://") or ref.startswith("https://"):
        return ref
    if not ref.startswith("/static/"):
        ref = f"/static/uploads/projects/{ref.lstrip('/')}"
    rel = ref.lstrip("/")
    path = BASE_DIR / rel.replace("/", os.sep)
    return ref if path.is_file() else ""


def _parse_hobbies_split() -> tuple[list[str], list[str]]:
    rows = _load_sheet_rows("爱好兴趣")
    hobbies: list[str] = []
    interests: list[str] = []
    if len(rows) < 2:
        return hobbies, interests
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = _cell_str(row[0])
        kind = _cell_str(row[1]) if len(row) > 1 else "爱好"
        if not name:
            continue
        if kind == "兴趣":
            interests.append(name)
        else:
            hobbies.append(name)
    return hobbies, interests


def _parse_school() -> dict:
    rows = _load_sheet_rows("学校组织")
    timeline: list[dict] = []
    if len(rows) < 2:
        return {"timeline": timeline, "headline": "", "education_line": ""}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = _cell_str(row[0])
        level = _cell_str(row[1]) if len(row) > 1 else ""
        if name:
            timeline.append({"name": name, "level": level})
    headline = ""
    for item in reversed(timeline):
        if item["level"] == "本科":
            headline = f"{item['name']} · {item['level']}"
            break
    if not headline and timeline:
        last = timeline[-1]
        headline = f"{last['name']} · {last['level']}" if last["level"] else last["name"]
    education_line = " → ".join(
        f"{t['name']}（{t['level']}）" if t["level"] else t["name"] for t in timeline[-3:]
    )
    return {"timeline": timeline, "headline": headline, "education_line": education_line}


def _parse_books() -> list[dict]:
    rows = _load_sheet_rows("书籍")
    if len(rows) < 2:
        return []
    books: list[dict] = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        title = _cell_str(row[0])
        if title:
            books.append({
                "title": title,
                "author": _cell_str(row[1]) if len(row) > 1 else "",
                "category": _cell_str(row[2]) if len(row) > 2 else "",
            })
    return books


def _parse_media_list(sheet: str) -> list[dict]:
    rows = _load_sheet_rows(sheet)
    if len(rows) < 2:
        return []
    items: list[dict] = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = _cell_str(row[0])
        if name:
            items.append({"name": name, "series": _cell_str(row[1]) if len(row) > 1 else ""})
    return items


def _parse_games() -> list[dict]:
    rows = _load_sheet_rows("游戏")
    if len(rows) < 2:
        return []
    items: list[dict] = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = _cell_str(row[0])
        if name:
            items.append({
                "name": name,
                "series": _cell_str(row[1]) if len(row) > 1 else "",
                "platform": _cell_str(row[2]) if len(row) > 2 else "",
            })
    return items


def _slug_from_name(name: str) -> str:
    s = name.lower()
    s = re.sub(r"[^\w\s-]", "", s, flags=re.UNICODE)
    s = re.sub(r"[\s_]+", "-", s).strip("-")
    return s or "project"


def _parse_projects_from_xlsx() -> list[dict]:
    rows = _load_sheet_rows("项目应用")
    if len(rows) < 2:
        return []
    projects: list[dict] = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = _cell_str(row[0])
        if not name:
            continue
        tech = _cell_str(row[5]) if len(row) > 5 else ""
        tags = (
            [t.strip() for t in tech.replace("，", ",").split(",") if t.strip()]
            if tech
            else []
        )
        feat_raw = _cell_str(row[6]) if len(row) > 6 else ""
        feature_lines = [h.strip() for h in feat_raw.split("\n") if h.strip()]
        github = _cell_str(row[4]) if len(row) > 4 else ""
        summary = _cell_str(row[3]) if len(row) > 3 else ""
        raw_img = _cell_str(row[7]) if len(row) > 7 else ""
        projects.append({
            "id": _slug_from_name(name),
            "title": name,
            "summary": summary,
            "project_type": _cell_str(row[1]) if len(row) > 1 else "",
            "category": _cell_str(row[2]) if len(row) > 2 else "项目",
            "year": "2025",
            "tags": tags[:8],
            "github": github or "#",
            "wiki_slug": _slug_from_name(name),
            "thumb": (len(projects) % 6) + 1,
            "highlights": feature_lines[:6] if feature_lines else [],
            "feature_lines": feature_lines,
            "image_url": raw_img,
            "cover_url": _resolve_project_image_url(raw_img),
            "repo_key": _github_repo_key(github),
        })
    return projects


def _parse_skills_from_projects(projects: list[dict]) -> list[dict]:
    freq: dict[str, int] = {}
    for proj in projects:
        for tag in proj.get("tags", []):
            if len(tag) <= 32:
                freq[tag] = freq.get(tag, 0) + 1
    ordered = sorted(freq.keys(), key=lambda k: (-freq[k], k))
    widths = ["92%", "90%", "88%", "85%", "82%", "80%"]
    if not ordered:
        return [
            {"name": "Python / FastAPI", "width": "92%"},
            {"name": "Android / Kotlin", "width": "88%"},
            {"name": "PyTorch / ONNX", "width": "85%"},
        ]
    return [
        {"name": name, "width": widths[i % len(widths)]}
        for i, name in enumerate(ordered[:6])
    ]


def _build_bio(school: dict, xlsx_projects: list[dict]) -> str:
    parts: list[str] = []
    if school.get("headline"):
        parts.append(f"{school['headline']}在读")
    parts.append("主攻 Python / Kotlin 与 Android，覆盖 CV、语音、OCR 与端侧部署。")
    if xlsx_projects:
        domains = []
        for p in xlsx_projects[:3]:
            d = p.get("category") or p.get("project_type", "")
            if d and d not in domains:
                domains.append(d.split("、")[0][:24])
        if domains:
            parts.append(f"近期项目：{'、'.join(domains)}。")
    return "".join(parts) if parts else DEFAULT_SITE["bio"]


def _build_focus(hobbies: list[str], interests: list[str], projects: list[dict]) -> str:
    chunks: list[str] = []
    if interests:
        chunks.append(" · ".join(interests[:3]))
    if hobbies:
        chunks.append(" · ".join(hobbies[:2]))
    if projects:
        chunks.append(projects[0].get("category", "")[:20])
    text = " · ".join(c for c in chunks if c)
    return text[:80] if text else DEFAULT_SITE["focus"]


def _build_hero_subtitle(projects: list[dict]) -> str:
    if not projects:
        return DEFAULT_SITE["hero_subtitle"]
    types = []
    for p in projects:
        t = p.get("project_type") or p.get("category", "")
        if t:
            short = t.split("/")[0].split("·")[0].strip()[:18]
            if short and short not in types:
                types.append(short)
    if types:
        return " · ".join(types[:4])
    return DEFAULT_SITE["hero_subtitle"]


def _parse_articles() -> list[dict]:
    rows = _load_sheet_rows("文章")
    if len(rows) < 2:
        return []
    items: list[dict] = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        title = _cell_str(row[0])
        if not title:
            continue
        category = _cell_str(row[1]) if len(row) > 1 else ""
        content = _cell_str(row[2]) if len(row) > 2 else ""
        excerpt = content[:300] + ("…" if len(content) > 300 else "")
        items.append({"title": title, "category": category, "excerpt": excerpt})
    return items


def build_data_hubs(data: dict, project_count: int = 0) -> list[dict]:
    """Homepage hub tiles — browse categories (projects use /projects under 站点)."""
    del project_count  # kept for call-site compatibility
    hobbies: list[str] = data.get("hobbies_list", [])
    interests: list[str] = data.get("interests_list", [])
    return [
        {
            "id": "hobbies",
            "label": "爱好兴趣",
            "sheet": "爱好兴趣",
            "count": len(hobbies) + len(interests),
            "url": "/browse/hobbies",
        },
        {
            "id": "books",
            "label": "书籍",
            "sheet": "书籍",
            "count": len(data.get("books", [])),
            "url": "/browse/books",
        },
        {
            "id": "anime",
            "label": "番剧",
            "sheet": "番剧",
            "count": len(data.get("anime", [])),
            "url": "/browse/anime",
        },
        {
            "id": "movies",
            "label": "电影",
            "sheet": "电影",
            "count": len(data.get("movies", [])),
            "url": "/browse/movies",
        },
        {
            "id": "games",
            "label": "游戏",
            "sheet": "游戏",
            "count": len(data.get("games", [])),
            "url": "/browse/games",
        },
        {
            "id": "school",
            "label": "教育经历",
            "sheet": "学校组织",
            "count": len(data.get("school_timeline", [])),
            "url": "/browse/school",
        },
    ]


DISABLED_HUB_IDS = frozenset({"certs", "hobbies", "school", "articles"})


def get_browse_page(hub_id: str) -> dict | None:
    """Build list/section payload for /browse/{hub_id}."""
    if hub_id in DISABLED_HUB_IDS:
        return None
    data = load_site_data()

    def _items_titled(rows: list[dict], title_key: str = "title", meta_fn=None) -> list[dict]:
        out: list[dict] = []
        for row in rows:
            title = row.get(title_key, "")
            if not title:
                continue
            meta = meta_fn(row) if meta_fn else row.get("meta", "")
            out.append({
                "title": title,
                "meta": meta,
                "image_url": row.get("image_url", ""),
                "excerpt": row.get("excerpt", ""),
            })
        return out

    if hub_id == "hobbies":
        hobbies = data.get("hobbies_list", [])
        interests = data.get("interests_list", [])
        return {
            "hub_id": hub_id,
            "page_title": "爱好与兴趣",
            "sheet_name": "爱好兴趣",
            "layout": "sections",
            "sections": [
                {
                    "title": "爱好",
                    "items": [{"title": n, "meta": "爱好", "image_url": "", "excerpt": ""} for n in hobbies],
                },
                {
                    "title": "兴趣",
                    "items": [{"title": n, "meta": "兴趣", "image_url": "", "excerpt": ""} for n in interests],
                },
            ],
        }
    if hub_id == "books":
        books = sorted(data.get("books", []), key=lambda b: (b.get("title") or "").lower())
        return {
            "hub_id": hub_id,
            "page_title": "书籍",
            "sheet_name": "书籍",
            "layout": "table",
            "table_columns": ["书名", "作者", "分类"],
            "items": [
                {
                    "title": b.get("title", ""),
                    "cells": [
                        b.get("title", ""),
                        b.get("author", "") or "—",
                        b.get("category", "") or "—",
                    ],
                    "meta": "",
                    "image_url": "",
                    "excerpt": "",
                }
                for b in books
                if b.get("title")
            ],
        }
    if hub_id == "anime":
        rows = sorted(data.get("anime", []), key=lambda r: (r.get("name") or "").lower())
        return {
            "hub_id": hub_id,
            "page_title": "番剧",
            "sheet_name": "番剧",
            "layout": "table",
            "table_columns": ["名称", "系列"],
            "items": [
                {
                    "title": r.get("name", ""),
                    "cells": [r.get("name", ""), r.get("series", "") or "—"],
                    "meta": "",
                    "image_url": "",
                    "excerpt": "",
                }
                for r in rows
                if r.get("name")
            ],
        }
    if hub_id == "movies":
        rows = sorted(data.get("movies", []), key=lambda r: (r.get("name") or "").lower())
        return {
            "hub_id": hub_id,
            "page_title": "电影",
            "sheet_name": "电影",
            "layout": "table",
            "table_columns": ["名称", "系列"],
            "items": [
                {
                    "title": r.get("name", ""),
                    "cells": [r.get("name", ""), r.get("series", "") or "—"],
                    "meta": "",
                    "image_url": "",
                    "excerpt": "",
                }
                for r in rows
                if r.get("name")
            ],
        }
    if hub_id == "games":
        rows = sorted(data.get("games", []), key=lambda r: (r.get("name") or "").lower())
        return {
            "hub_id": hub_id,
            "page_title": "游戏",
            "sheet_name": "游戏",
            "layout": "table",
            "table_columns": ["名称", "系列", "平台"],
            "items": [
                {
                    "title": r.get("name", ""),
                    "cells": [
                        r.get("name", ""),
                        r.get("series", "") or "—",
                        r.get("platform", "") or "—",
                    ],
                    "meta": "",
                    "image_url": "",
                    "excerpt": "",
                }
                for r in rows
                if r.get("name")
            ],
        }
    if hub_id == "school":
        timeline = data.get("school_timeline", [])
        return {
            "hub_id": hub_id,
            "page_title": "教育经历",
            "sheet_name": "学校组织",
            "layout": "list",
            "items": [
                {
                    "title": t["name"],
                    "meta": t.get("level", ""),
                    "image_url": "",
                    "excerpt": "",
                }
                for t in timeline
            ],
        }
    return None


def merge_projects_catalog(base_projects: list[dict], xlsx_projects: list[dict] | None = None) -> list[dict]:
    """Enrich wiki-backed projects with spreadsheet fields; keep ids/slugs from base."""
    if xlsx_projects is None:
        xlsx_projects = _parse_projects_from_xlsx()
    by_repo = {p["repo_key"]: p for p in xlsx_projects if p.get("repo_key")}
    merged: list[dict] = []

    for base in base_projects:
        p = dict(base)
        key = _github_repo_key(p.get("github", ""))
        x = by_repo.get(key) if key else None
        if x:
            if x.get("summary"):
                p["summary"] = x["summary"]
            if x.get("tags"):
                p["tags"] = x["tags"]
            if x.get("highlights"):
                hl = list(x["highlights"])
                summary = (p.get("summary") or "").strip()
                p["highlights"] = [h for h in hl if h.strip() and h.strip() != summary]
            if x.get("feature_lines"):
                fl = list(x["feature_lines"])
                summary = (p.get("summary") or "").strip()
                p["feature_lines"] = [h for h in fl if h.strip() and h.strip() != summary]
            if x.get("category"):
                p["category"] = x["category"]
            if x.get("project_type"):
                p["project_type"] = x["project_type"]
            if x.get("image_url"):
                p["image_url"] = x["image_url"]
                p["cover_url"] = _resolve_project_image_url(x["image_url"])
        if not p.get("cover_url") and p.get("image_url"):
            p["cover_url"] = _resolve_project_image_url(p["image_url"])
        if not p.get("feature_lines") and p.get("highlights"):
            p["feature_lines"] = list(p["highlights"])
        merged.append(p)

    base_keys = {_github_repo_key(p.get("github", "")) for p in base_projects if p.get("github")}
    for x in xlsx_projects:
        rk = x.get("repo_key")
        if rk and rk in base_keys:
            continue
        merged.append(x)
    return merged


def clear_site_data_cache() -> None:
    """Call after zhita_settings.xlsx changes (or restart the server)."""
    load_site_data.cache_clear()
    try:
        from app.projects import clear_projects_cache
        from app.context import clear_context_cache
        from app.markdown.page_cache import clear_render_cache

        clear_projects_cache()
        clear_context_cache()
        clear_render_cache()
    except ImportError:
        pass


@lru_cache(maxsize=1)
def load_site_data() -> dict:
    identity = _load_identity_json()
    ui = dict(DEFAULT_SITE["ui"])
    site_name = identity["site_name"]
    site_title = identity["site_title"]

    hobbies, interests = _parse_hobbies_split()
    school = _parse_school()
    books = _parse_books()
    anime = _parse_media_list("番剧")
    movies = _parse_media_list("电影")
    games = _parse_games()
    articles = _parse_articles()
    xlsx_projects = _parse_projects_from_xlsx()

    skills = _parse_skills_from_projects(xlsx_projects)
    tech_tags = [t for p in xlsx_projects for t in p.get("tags", [])]
    seen: set[str] = set()
    tech_tags_unique: list[str] = []
    for t in tech_tags:
        if t and t not in seen and len(t) <= 28:
            seen.add(t)
            tech_tags_unique.append(t)

    hobby_tags = hobbies + interests
    media_count = len(books) + len(anime) + len(movies)

    payload = {
        "site_name": site_name,
        "site_title": site_title,
        "tagline": DEFAULT_SITE["tagline"],
        "hero_subtitle": _build_hero_subtitle(xlsx_projects),
        "bio": _build_bio(school, xlsx_projects),
        "focus": _build_focus(hobbies, interests, xlsx_projects),
        "availability": school.get("headline") or DEFAULT_SITE["availability"],
        "education_line": school.get("education_line", ""),
        "school_timeline": school.get("timeline", []),
        "ui": ui,
        "skills": skills,
        "hobby_tags": hobby_tags,
        "tech_tags": tech_tags_unique[:14],
        "books": books,
        "books_featured": books[:6],
        "anime": anime,
        "movies": movies,
        "games": games,
        "games_count": len(games),
        "articles": articles,
        "hobbies_list": hobbies,
        "interests_list": interests,
        "xlsx_projects": xlsx_projects,
        "reading_line": f"在读 / 读过 {len(books)} 册 · 番剧 {len(anime)} · 电影 {len(movies)} · Steam {len(games)}",
    }
    hub_count = len(build_data_hubs(payload, len(xlsx_projects)))
    payload["stats"] = [
        {"value": len(xlsx_projects), "unit": ui["stat_projects_unit"], "label": ui["stat_projects"]},
        {"value": hub_count, "unit": ui["stat_skills_unit"], "label": ui["stat_skills"]},
        {"value": media_count, "unit": ui["stat_media_unit"], "label": ui["stat_media"]},
    ]
    return payload


def get_spreadsheet_context() -> dict:
    data = load_site_data()
    return {
        "site_name": data["site_name"],
        "site_title": data["site_title"],
        "tagline": data["tagline"],
        "hero_subtitle": data["hero_subtitle"],
        "bio": data["bio"],
        "focus": data["focus"],
        "availability": data["availability"],
        "education_line": data["education_line"],
        "reading_line": data["reading_line"],
        "ui": data["ui"],
        "skills": data["skills"],
        "hobby_tags": data["hobby_tags"],
        "tech_tags": data["tech_tags"],
        "stats": data["stats"],
        "books_featured": data["books_featured"],
        "games_count": data["games_count"],
    }
